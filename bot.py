import os
import io
import logging
import pyotp
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
from telegram import Update, BotCommand
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

from database import Database

load_dotenv()

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
DB_PATH = os.getenv("DB_PATH", "./data/dashboard.db")
ADMIN_IDS = [int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()]

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

db = Database(DB_PATH)


# --- Helpers ---

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


async def is_admin_full(user_id: int) -> bool:
    """Check .env admins AND database admins."""
    if user_id in ADMIN_IDS:
        return True
    return await db.is_db_admin(user_id)


def generate_otp(secret: str) -> tuple[str, int]:
    """Returns (otp_code, seconds_remaining)."""
    try:
        totp = pyotp.TOTP(secret)
        code = totp.now()
        remaining = 30 - (int(datetime.utcnow().timestamp()) % 30)
        return code, remaining
    except Exception:
        return "ERROR", 0


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse Excel file and return list of account dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    headers = []
    accounts = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip().lower() if h else "" for h in row]
            continue
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))

        account_id = str(row_dict.get("id") or row_dict.get("username") or row_dict.get("email") or "").strip()
        password = str(row_dict.get("password") or "").strip()
        totp_secret = str(row_dict.get("totp secret") or row_dict.get("totp_secret") or row_dict.get("totp") or "").strip()

        if account_id and password and totp_secret:
            accounts.append({
                "id": account_id,
                "password": password,
                "totp_secret": totp_secret
            })

    return accounts


def parse_users_excel(file_bytes: bytes) -> list[dict]:
    """Parse Excel file with Phone and Name columns."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = []
    users = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip().lower() if h else "" for h in row]
            continue
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        phone = str(row_dict.get("phone") or row_dict.get("mobile") or row_dict.get("number") or "").strip()
        name = str(row_dict.get("name") or row_dict.get("full name") or row_dict.get("fullname") or "").strip()
        if phone and name:
            users.append({"phone": phone, "name": name})
    return users


# --- Admin Commands ---

async def cmd_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    await update.message.reply_text(
        "📤 *Upload your Excel file now.*\n\n"
        "The file must have these columns:\n"
        "`ID` | `Password` | `TOTP Secret` | `Status` (optional)\n\n"
        "Send the `.xlsx` file as a document.",
        parse_mode="Markdown"
    )
    context.user_data["waiting_for_upload"] = True


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        return

    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Please upload a `.xlsx` file.")
        return

    # Determine which upload mode
    waiting_accounts = context.user_data.get("waiting_for_upload")
    waiting_users = context.user_data.get("waiting_for_users_upload")

    if not waiting_accounts and not waiting_users:
        return

    await update.message.reply_text("⏳ Processing file...")

    try:
        file = await doc.get_file()
        file_bytes = await file.download_as_bytearray()
    except Exception as e:
        await update.message.reply_text(f"❌ Failed to download file: {e}")
        return

    if waiting_accounts:
        # Parse accounts Excel
        try:
            accounts = parse_excel(bytes(file_bytes))
        except Exception as e:
            await update.message.reply_text(f"❌ Failed to parse file: {e}")
            return

        if not accounts:
            await update.message.reply_text(
                "❌ No valid accounts found.\n"
                "Make sure columns are: `ID`, `Password`, `TOTP Secret`"
            )
            return

        added = updated = 0
        existing_ids = [a["account_id"] for a in await db.get_all_accounts()]
        for acc in accounts:
            if acc["id"] in existing_ids:
                updated += 1
            else:
                added += 1
            await db.upsert_account(acc["id"], acc["password"], acc["totp_secret"])

        context.user_data["waiting_for_upload"] = False
        stats = await db.get_stats()
        await update.message.reply_text(
            f"✅ *Accounts Upload Complete!*\n\n"
            f"📥 Added: {added} | 🔄 Updated: {updated}\n"
            f"📊 Total: {stats['total']} | Available: {stats['available']} | Assigned: {stats['assigned']}",
            parse_mode="Markdown"
        )

    elif waiting_users:
        # Parse users Excel
        try:
            users = parse_users_excel(bytes(file_bytes))
        except Exception as e:
            await update.message.reply_text(f"❌ Failed to parse file: {e}")
            return

        if not users:
            await update.message.reply_text(
                "❌ No valid users found.\n"
                "Make sure columns are: `Phone`, `Name`"
            )
            return

        for u in users:
            await db.upsert_allowed_user(u["phone"], u["name"])

        context.user_data["waiting_for_users_upload"] = False
        await update.message.reply_text(
            f"✅ *User List Upload Complete!*\n\n"
            f"👥 {len(users)} users added/updated in whitelist.\n"
            f"Use /listusers to view them.",
            parse_mode="Markdown"
        )


async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    accounts = await db.get_all_accounts()
    if not accounts:
        await update.message.reply_text("📭 No accounts in the database.")
        return

    stats = await db.get_stats()

    # Send in chunks to avoid message length limits
    chunk_size = 20
    chunks = [accounts[i:i+chunk_size] for i in range(0, len(accounts), chunk_size)]

    for idx, chunk in enumerate(chunks):
        lines = []
        for acc in chunk:
            status_emoji = "🟢" if acc["status"] == "available" else "🔴"
            assigned = f"(User: {acc['assigned_to']})" if acc.get("assigned_to") else ""
            lines.append(f"{status_emoji} `{acc['account_id']}` {assigned}")

        header = f"📋 *Account List ({idx+1}/{len(chunks)})*\n\n" if idx == 0 else ""
        await update.message.reply_text(
            f"{header}" + "\n".join(lines),
            parse_mode="Markdown"
        )

    await update.message.reply_text(
        f"📊 *Summary:* Total: {stats['total']} | 🟢 Available: {stats['available']} | 🔴 Assigned: {stats['assigned']}",
        parse_mode="Markdown"
    )


async def cmd_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    await db.reset_all()
    stats = await db.get_stats()
    await update.message.reply_text(
        f"✅ All {stats['total']} accounts reset to *available*.",
        parse_mode="Markdown"
    )


async def cmd_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/remove <account_id>`", parse_mode="Markdown")
        return

    account_id = " ".join(context.args).strip()
    success = await db.remove_account(account_id)
    if success:
        await update.message.reply_text(f"✅ Account `{account_id}` removed.", parse_mode="Markdown")
    else:
        await update.message.reply_text(f"❌ Account `{account_id}` not found.", parse_mode="Markdown")


async def cmd_resetaccount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/resetaccount <account_id>`", parse_mode="Markdown")
        return
    account_id = " ".join(context.args).strip()
    await db.reset_account(account_id)
    await update.message.reply_text(f"✅ Account `{account_id}` reset to available.", parse_mode="Markdown")


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    stats = await db.get_stats()
    await update.message.reply_text(
        f"📊 *Account Statistics*\n\n"
        f"Total: `{stats['total']}`\n"
        f"🟢 Available: `{stats['available']}`\n"
        f"🔴 Assigned: `{stats['assigned']}`",
        parse_mode="Markdown"
    )


# --- User Commands ---

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin_notice = "\n\n👑 *You are an admin.* Use /upload to add accounts." if await is_admin_full(update.effective_user.id) else ""
    await update.message.reply_text(
        f"👋 *Welcome to Dashboard Account Bot!*\n\n"
        f"Use /getaccount to receive an available account.\n"
        f"Use /myotp to refresh your current OTP code.{admin_notice}",
        parse_mode="Markdown"
    )


async def cmd_getaccount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Check if user already has an account
    existing = await db.get_user_account(user_id)
    if existing:
        otp, remaining = generate_otp(existing["totp_secret"])
        await update.message.reply_text(
            f"ℹ️ You already have an assigned account.\n\n"
            f"🆔 *Account ID:* `{existing['account_id']}`\n"
            f"🔑 *Password:* `{existing['password']}`\n"
            f"🔐 *OTP Code:* `{otp}`\n"
            f"⏱ *Expires in:* {remaining}s\n\n"
            f"Use /myotp to get a fresh OTP anytime.",
            parse_mode="Markdown"
        )
        return

    # Check if an account is even available before asking for number
    account = await db.get_available_account()
    if not account:
        await update.message.reply_text(
            "❌ *No accounts available right now.*\n"
            "Please contact an admin.",
            parse_mode="Markdown"
        )
        return

    # Ask for mobile number first
    context.user_data["waiting_for_mobile"] = True
    await update.message.reply_text(
        "📱 *Please enter your mobile number* to proceed:\n\n"
        "Example: `+91 9876543210`",
        parse_mode="Markdown"
    )


async def handle_mobile_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle mobile number submitted before account assignment."""
    if not context.user_data.get("waiting_for_mobile"):
        return

    user_id = update.effective_user.id
    mobile = update.message.text.strip()

    # Basic validation — must have at least 7 digits
    digits = ''.join(c for c in mobile if c.isdigit())
    if len(digits) < 7:
        await update.message.reply_text(
            "❌ Invalid mobile number. Please enter a valid number.\n"
            "Example: `+91 9876543210`",
            parse_mode="Markdown"
        )
        return

    # Check whitelist
    allowed = await db.check_allowed_phone(mobile)
    if allowed is None:
        await update.message.reply_text(
            "❌ *Your number is not registered.*\n\n"
            "Only registered users can receive accounts.\n"
            "Please contact admin via /adcm.",
            parse_mode="Markdown"
        )
        return

    user_name = allowed["name"]
    context.user_data["waiting_for_mobile"] = False

    # Get available account
    account = await db.get_available_account()
    if not account:
        await update.message.reply_text(
            "❌ *No accounts available right now.*\n"
            "Please contact an admin.",
            parse_mode="Markdown"
        )
        return

    # Assign it
    await db.assign_account(account["account_id"], user_id)
    await db.mark_phone_used(mobile)
    otp, remaining = generate_otp(account["totp_secret"])

    user = update.effective_user
    username = f"@{user.username}" if user.username else f"ID: {user_id}"

    # Send account details to user
    await update.message.reply_text(
        f"✅ *Welcome, {user_name}!*\n\n"
        f"*Account Assigned to You:*\n\n"
        f"🆔 *Account ID:* `{account['account_id']}`\n"
        f"🔑 *Password:* `{account['password']}`\n"
        f"🔐 *OTP Code:* `{otp}`\n"
        f"⏱ *OTP Expires in:* {remaining}s\n\n"
        f"⚠️ This is your *one-time OTP*. Save it now.\n"
        f"For further OTP assistance, contact admin via /adcm.",
        parse_mode="Markdown"
    )

    # Mark OTP as shown
    await db.mark_otp_shown(account["account_id"])

    # Notify all admins / group
    notify_group_id = await db.get_notify_group_chat_id()
    notification_text = (
        f"📋 *New Account Assignment*\n\n"
        f"👤 User: {username} (`{user_id}`)\n"
        f"📝 Name: `{user_name}`\n"
        f"📱 Mobile: `{mobile}`\n"
        f"🆔 Account: `{account['account_id']}`\n"
        f"🔑 Password: `{account['password']}`\n"
        f"⏰ Assigned: just now"
    )

    sent = False
    if notify_group_id:
        try:
            await context.bot.send_message(chat_id=notify_group_id, text=notification_text, parse_mode="Markdown")
            sent = True
        except Exception:
            pass

    if not sent:
        all_admin_ids = list(ADMIN_IDS)
        db_admins = await db.get_admins()
        for a in db_admins:
            if a["user_id"] not in all_admin_ids:
                all_admin_ids.append(a["user_id"])
        for admin_id in all_admin_ids:
            try:
                await context.bot.send_message(chat_id=admin_id, text=notification_text, parse_mode="Markdown")
            except Exception:
                pass

    logger.info(f"Account {account['account_id']} assigned to {user_name} ({mobile})")


async def cmd_myotp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    account = await db.get_user_account(user_id)

    if not account:
        await update.message.reply_text(
            "❌ You don't have an assigned account.\n"
            "Use /getaccount to get one.",
            parse_mode="Markdown"
        )
        return

    # Block if OTP already shown once
    if account.get("otp_shown", 0):
        contact = await db.get_admin_contact()
        contact_text = ""
        if contact:
            if contact.get("telegram"):
                contact_text += f"\n📱 Telegram: {contact['telegram']}"
            if contact.get("phone"):
                contact_text += f"\n📞 Phone: `{contact['phone']}`"
            if contact.get("note"):
                contact_text += f"\nℹ️ {contact['note']}"
        await update.message.reply_text(
            f"⚠️ *OTP already provided once.*\n\n"
            f"For security, OTP can only be shown once per account.\n"
            f"Please contact admin for further assistance.{contact_text}",
            parse_mode="Markdown"
        )
        return

    otp, remaining = generate_otp(account["totp_secret"])
    await db.mark_otp_shown(account["account_id"])
    await update.message.reply_text(
        f"🔐 *Your OTP Code*\n\n"
        f"🆔 Account: `{account['account_id']}`\n"
        f"🔢 *Code:* `{otp}`\n"
        f"⏱ *Valid for:* {remaining} seconds\n\n"
        f"⚠️ This is your *last OTP*. Contact admin for more help.",
        parse_mode="Markdown"
    )


async def cmd_myaccount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    account = await db.get_user_account(user_id)

    if not account:
        await update.message.reply_text(
            "❌ You don't have an assigned account.\n"
            "Use /getaccount to get one."
        )
        return

    await update.message.reply_text(
        f"📋 *Your Account Details*\n\n"
        f"🆔 *Account ID:* `{account['account_id']}`\n"
        f"🔑 *Password:* `{account['password']}`\n"
        f"📊 *Status:* `{account['status']}`",
        parse_mode="Markdown"
    )


async def cmd_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """User signals they are done and want to withdraw."""
    user_id = update.effective_user.id
    account = await db.get_user_account(user_id)

    if not account:
        await update.message.reply_text(
            "❌ You don't have an assigned account.\n"
            "Use /getaccount to get one first."
        )
        return

    # Check if already has a pending withdrawal
    existing = await db.get_user_withdrawal(user_id)
    if existing:
        await update.message.reply_text(
            "⏳ You already have a pending withdrawal request.\n"
            "Please wait for admin to process it.\n\n"
            f"Request ID: `#{existing['id']}`",
            parse_mode="Markdown"
        )
        return

    # Get the active withdrawal wallet
    wallet = await db.get_active_wallet()
    if not wallet:
        await update.message.reply_text(
            "❌ No withdrawal wallet configured yet.\n"
            "Please contact an admin."
        )
        return

    # Create withdrawal record
    withdrawal_id = await db.create_withdrawal(user_id, account["account_id"])

    # Ask for screenshot
    context.user_data["pending_withdrawal_id"] = withdrawal_id
    context.user_data["waiting_for_screenshot"] = True

    await update.message.reply_text(
        f"✅ *Withdrawal Request Started*\n\n"
        f"Please withdraw your earnings to this wallet:\n\n"
        f"🌐 *Network:* `{wallet['network']}`\n"
        f"💳 *Wallet Address:*\n`{wallet['address']}`\n\n"
        f"📸 After sending, *send a screenshot* of your withdrawal confirmation so admin can verify.",
        parse_mode="Markdown"
    )


async def handle_screenshot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle screenshot submitted by user after /done."""
    if not context.user_data.get("waiting_for_screenshot"):
        return

    user_id = update.effective_user.id
    withdrawal_id = context.user_data.get("pending_withdrawal_id")

    if not withdrawal_id:
        return

    # Must be a photo
    if not update.message.photo:
        await update.message.reply_text(
            "📸 Please send a *screenshot* (photo) of your withdrawal confirmation.",
            parse_mode="Markdown"
        )
        return

    # Get highest quality photo
    file_id = update.message.photo[-1].file_id
    caption = update.message.caption or ""

    await db.update_withdrawal_screenshot(withdrawal_id, file_id)
    if caption:
        await db.update_withdrawal_txn(withdrawal_id, caption)

    context.user_data["waiting_for_screenshot"] = False
    context.user_data["pending_withdrawal_id"] = None

    account = await db.get_user_account(user_id)
    user = update.effective_user
    username = f"@{user.username}" if user.username else f"ID: {user_id}"

    # Notify group if set, else notify all admins
    notify_group_id = await db.get_notify_group_chat_id()
    notification_text = (
        f"🔔 *New Withdrawal Request*\n\n"
        f"Request ID: `#{withdrawal_id}`\n"
        f"👤 User: {username} (`{user_id}`)\n"
        f"🆔 Account: `{account['account_id'] if account else 'N/A'}`\n"
        f"📝 Caption: {caption or 'None'}\n\n"
        f"Use:\n"
        f"`/approve {withdrawal_id}` — Approve & free account\n"
        f"`/reject {withdrawal_id} [reason]` — Reject request"
    )

    sent_to_group = False
    if notify_group_id:
        try:
            await context.bot.send_photo(
                chat_id=notify_group_id,
                photo=file_id,
                caption=notification_text,
                parse_mode="Markdown"
            )
            sent_to_group = True
        except Exception as e:
            logger.warning(f"Could not send to group {notify_group_id}: {e}")

    if not sent_to_group:
        # Fallback to all admins
        all_admin_ids = list(ADMIN_IDS)
        db_admins = await db.get_admins()
        for a in db_admins:
            if a["user_id"] not in all_admin_ids:
                all_admin_ids.append(a["user_id"])

        for admin_id in all_admin_ids:
            try:
                await context.bot.send_photo(
                    chat_id=admin_id,
                    photo=file_id,
                    caption=notification_text,
                    parse_mode="Markdown"
                )
            except Exception:
                pass

    await update.message.reply_text(
        f"✅ *Withdrawal request submitted!*\n\n"
        f"Request ID: `#{withdrawal_id}`\n\n"
        f"Admin will verify your screenshot and process the request shortly.\n"
        f"Your account will be released after approval.",
        parse_mode="Markdown"
    )


async def cmd_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin approves a withdrawal and frees the account."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/approve <withdrawal_id> [note]`", parse_mode="Markdown")
        return
    try:
        wid = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid withdrawal ID.")
        return

    note = " ".join(context.args[1:]) if len(context.args) > 1 else None
    withdrawal = await db.get_withdrawal_by_id(wid)
    if not withdrawal:
        await update.message.reply_text(f"❌ Withdrawal `#{wid}` not found.", parse_mode="Markdown")
        return
    if withdrawal["status"] != "pending":
        await update.message.reply_text(f"❌ Withdrawal `#{wid}` is already `{withdrawal['status']}`.", parse_mode="Markdown")
        return

    await db.approve_withdrawal(wid, update.effective_user.id, note)
    # Free the account
    await db.reset_account(withdrawal["account_id"])

    await update.message.reply_text(
        f"✅ Withdrawal `#{wid}` *approved*.\n"
        f"Account `{withdrawal['account_id']}` is now available again.",
        parse_mode="Markdown"
    )

    # Notify user
    try:
        msg = f"✅ *Your withdrawal has been approved!*\n\nRequest ID: `#{wid}`\nYour account has been released."
        if note:
            msg += f"\nAdmin note: {note}"
        await context.bot.send_message(chat_id=withdrawal["user_id"], text=msg, parse_mode="Markdown")
    except Exception:
        pass


async def cmd_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin rejects a withdrawal."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/reject <withdrawal_id> [reason]`", parse_mode="Markdown")
        return
    try:
        wid = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid withdrawal ID.")
        return

    reason = " ".join(context.args[1:]) if len(context.args) > 1 else "No reason provided"
    withdrawal = await db.get_withdrawal_by_id(wid)
    if not withdrawal:
        await update.message.reply_text(f"❌ Withdrawal `#{wid}` not found.", parse_mode="Markdown")
        return

    await db.reject_withdrawal(wid, update.effective_user.id, reason)

    await update.message.reply_text(
        f"❌ Withdrawal `#{wid}` *rejected*.\nReason: {reason}",
        parse_mode="Markdown"
    )

    # Notify user
    try:
        await context.bot.send_message(
            chat_id=withdrawal["user_id"],
            text=f"❌ *Your withdrawal was rejected.*\n\nRequest ID: `#{wid}`\nReason: {reason}\n\nPlease contact admin.",
            parse_mode="Markdown"
        )
    except Exception:
        pass


async def cmd_withdrawals(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin views all pending withdrawals."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    pending = await db.get_pending_withdrawals()
    if not pending:
        await update.message.reply_text("✅ No pending withdrawals.")
        return

    lines = []
    for w in pending:
        lines.append(
            f"📌 *#{w['id']}* — User `{w['user_id']}`\n"
            f"  Account: `{w['account_id']}`\n"
            f"  TXN: `{w.get('txn_id') or 'Not provided yet'}`\n"
            f"  Requested: {str(w['requested_at'])[:16]}"
        )
    await update.message.reply_text(
        f"📋 *Pending Withdrawals ({len(pending)})*\n\n" + "\n\n".join(lines),
        parse_mode="Markdown"
    )


async def cmd_setwallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin sets the withdrawal wallet address."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Usage: `/setwallet <address> <network> [label]`\n\n"
            "Example:\n`/setwallet TRxxxxxxxxxxxxxxxx USDT_TRC20 MainWallet`",
            parse_mode="Markdown"
        )
        return

    address = context.args[0]
    network = context.args[1]
    label = " ".join(context.args[2:]) if len(context.args) > 2 else "Main Wallet"

    wid = await db.add_wallet(label, address, network)
    await update.message.reply_text(
        f"✅ Wallet added (ID: `{wid}`)\n\n"
        f"Label: {label}\nNetwork: `{network}`\nAddress: `{address}`",
        parse_mode="Markdown"
    )


async def cmd_wallets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin views all withdrawal wallets."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    wallets = await db.get_all_wallets()
    if not wallets:
        await update.message.reply_text("No wallets set. Use `/setwallet` to add one.", parse_mode="Markdown")
        return

    lines = []
    for w in wallets:
        active = "✅ Active" if w["active"] else "❌ Inactive"
        lines.append(f"*#{w['id']} — {w['label']}* {active}\nNetwork: `{w['network']}`\nAddress: `{w['address']}`")

    await update.message.reply_text(
        "💳 *Withdrawal Wallets*\n\n" + "\n\n".join(lines),
        parse_mode="Markdown"
    )


async def cmd_uploadusers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin uploads Excel with allowed phone numbers and names."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    await update.message.reply_text(
        "📤 *Upload your user list Excel file.*\n\n"
        "The file must have these columns:\n"
        "`Phone` | `Name`\n\n"
        "Example:\n"
        "`+91 9876543210` | `Rahul Sharma`\n\n"
        "Send the `.xlsx` file now.",
        parse_mode="Markdown"
    )
    context.user_data["waiting_for_users_upload"] = True


async def cmd_listusers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin views all allowed users."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    users = await db.get_all_allowed_users()
    if not users:
        await update.message.reply_text(
            "📭 No users in whitelist.\n"
            "Use /uploadusers to upload a list."
        )
        return

    chunk_size = 30
    chunks = [users[i:i+chunk_size] for i in range(0, len(users), chunk_size)]
    for idx, chunk in enumerate(chunks):
        lines = []
        for u in chunk:
            used = "✅" if u["used"] else "⬜"
            lines.append(f"{used} `{u['phone']}` — {u['name']}")
        header = f"👥 *Allowed Users ({len(users)} total)*\n\n" if idx == 0 else ""
        await update.message.reply_text(header + "\n".join(lines), parse_mode="Markdown")


async def cmd_adduser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin adds a single user. Usage: /adduser <phone> <name>"""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Usage: `/adduser <phone> <name>`\n"
            "Example: `/adduser +919876543210 Rahul Sharma`",
            parse_mode="Markdown"
        )
        return
    phone = context.args[0]
    name = " ".join(context.args[1:])
    await db.upsert_allowed_user(phone, name)
    await update.message.reply_text(f"✅ User added:\n📱 `{phone}` — {name}", parse_mode="Markdown")


async def cmd_removeuser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin removes a user from whitelist. Usage: /removeuser <phone>"""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/removeuser <phone>`", parse_mode="Markdown")
        return
    phone = context.args[0]
    success = await db.remove_allowed_user(phone)
    if success:
        await update.message.reply_text(f"✅ Removed `{phone}` from whitelist.", parse_mode="Markdown")
    else:
        await update.message.reply_text(f"❌ `{phone}` not found in whitelist.", parse_mode="Markdown")


async def cmd_clearusers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin clears all allowed users."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    await db.clear_allowed_users()
    await update.message.reply_text("✅ All users cleared from whitelist.")


async def cmd_adotp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin gets OTP for any account by ID. Usage: /adotp <account_id>"""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text(
            "Usage: `/adotp <account_id>`\n\n"
            "Example: `/adotp user@example.com`",
            parse_mode="Markdown"
        )
        return

    account_id = " ".join(context.args).strip()

    # Fetch account from DB
    import aiosqlite as _aiosqlite
    async with _aiosqlite.connect(db.db_path) as conn:
        conn.row_factory = _aiosqlite.Row
        cursor = await conn.execute(
            "SELECT * FROM accounts WHERE account_id = ?", (account_id,)
        )
        row = await cursor.fetchone()

    if not row:
        await update.message.reply_text(
            f"❌ Account `{account_id}` not found in database.",
            parse_mode="Markdown"
        )
        return

    account = dict(row)
    otp, remaining = generate_otp(account["totp_secret"])

    await update.message.reply_text(
        f"🔐 *OTP for Account*\n\n"
        f"🆔 Account: `{account['account_id']}`\n"
        f"🔢 *OTP Code:* `{otp}`\n"
        f"⏱ *Valid for:* {remaining} seconds\n"
        f"📊 Status: `{account['status']}`\n"
        f"👤 Assigned to: `{account.get('assigned_to') or 'Nobody'}`",
        parse_mode="Markdown"
    )


async def cmd_addadmin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text(
            "Usage: `/addadmin <user_id>`\n\n"
            "To get a user's ID, ask them to forward a message to @userinfobot",
            parse_mode="Markdown"
        )
        return
    try:
        new_admin_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Must be a number.")
        return

    await db.add_admin(new_admin_id, update.effective_user.id)

    # Update their command list in Telegram
    from telegram import BotCommandScopeChat
    admin_cmds = [
        BotCommand("start", "Welcome message"),
        BotCommand("getaccount", "Get an available account with OTP"),
        BotCommand("myotp", "Refresh your current OTP code"),
        BotCommand("myaccount", "View your account details"),
        BotCommand("help", "Show all commands"),
        BotCommand("upload", "Upload Excel file with accounts"),
        BotCommand("list", "View all accounts and their status"),
        BotCommand("stats", "View account statistics"),
        BotCommand("reset", "Mark all accounts as available"),
        BotCommand("resetaccount", "Reset a specific account"),
        BotCommand("remove", "Remove an account permanently"),
        BotCommand("addadmin", "Add a new admin"),
        BotCommand("removeadmin", "Remove an admin"),
        BotCommand("admins", "List all admins"),
    ]
    try:
        await context.bot.set_my_commands(
            admin_cmds,
            scope=BotCommandScopeChat(chat_id=new_admin_id)
        )
    except Exception:
        pass

    await update.message.reply_text(
        f"✅ User `{new_admin_id}` added as admin.\n"
        f"They will see admin commands next time they open the bot.",
        parse_mode="Markdown"
    )


async def cmd_removeadmin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return
    if not context.args:
        await update.message.reply_text("Usage: `/removeadmin <user_id>`", parse_mode="Markdown")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return

    # Prevent removing .env admins
    if target_id in ADMIN_IDS:
        await update.message.reply_text("❌ Cannot remove admins set in `.env` from here. Remove them from ADMIN_IDS instead.")
        return

    success = await db.remove_admin(target_id)

    # Restore user-only command list
    from telegram import BotCommandScopeChat
    user_cmds = [
        BotCommand("start", "Welcome message"),
        BotCommand("getaccount", "Get an available account with OTP"),
        BotCommand("myotp", "Refresh your current OTP code"),
        BotCommand("myaccount", "View your account details"),
        BotCommand("help", "Show all commands"),
    ]
    try:
        await context.bot.set_my_commands(
            user_cmds,
            scope=BotCommandScopeChat(chat_id=target_id)
        )
    except Exception:
        pass

    if success:
        await update.message.reply_text(f"✅ User `{target_id}` removed from admins.", parse_mode="Markdown")
    else:
        await update.message.reply_text(f"❌ User `{target_id}` is not a database admin.", parse_mode="Markdown")


async def cmd_admins(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    admins = await db.get_admins()
    lines = []

    # .env admins
    for uid in ADMIN_IDS:
        lines.append(f"👑 `{uid}` — from .env (permanent)")

    # DB admins
    for a in admins:
        if a["user_id"] not in ADMIN_IDS:
            lines.append(f"🔑 `{a['user_id']}` — added by `{a['added_by']}` on {str(a['added_at'])[:10]}")

    if not lines:
        await update.message.reply_text("No admins configured.")
        return

    embed_text = "👑 *Admin List*\n\n" + "\n".join(lines)
    await update.message.reply_text(embed_text, parse_mode="Markdown")


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_commands = (
        "👤 *User Commands*\n"
        "/getaccount — Get an available account with OTP\n"
        "/myotp — Refresh your current OTP code\n"
        "/myaccount — View your account details\n"
        "/done — Signal you're done & request withdrawal\n"
        "/help — Show this message\n"
        "/start — Welcome message"
    )
    admin_commands = (
        "\n\n👑 *Admin Commands*\n"
        "/upload — Upload Excel file with accounts\n"
        "/list — View all accounts and their status\n"
        "/stats — View account statistics\n"
        "/reset — Mark all accounts as available\n"
        "/resetaccount <id> — Reset a specific account\n"
        "/remove <id> — Remove an account permanently\n"
        "/withdrawals — View all pending withdrawals\n"
        "/approve <id> [note] — Approve a withdrawal\n"
        "/reject <id> [reason] — Reject a withdrawal\n"
        "/setwallet <address> <network> [label] — Set wallet\n"
        "/wallets — View all withdrawal wallets\n"
        "/addadmin <user\\_id> — Add a new admin\n"
        "/removeadmin <user\\_id> — Remove an admin\n"
        "/admins — List all admins"
    )

    text = user_commands
    if await is_admin_full(update.effective_user.id):
        text += admin_commands

    await update.message.reply_text(text, parse_mode="Markdown")


async def cmd_adcm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin sets contact info. Usage: /adcm"""
    if not await is_admin_full(update.effective_user.id):
        # Show contact info to users
        contact = await db.get_admin_contact()
        if not contact:
            await update.message.reply_text("ℹ️ Admin contact not set yet. Please try later.")
            return
        text = "📞 *Contact Admin*\n"
        if contact.get("telegram"):
            text += f"\n📱 Telegram: {contact['telegram']}"
        if contact.get("phone"):
            text += f"\n☎️ Phone: `{contact['phone']}`"
        if contact.get("note"):
            text += f"\nℹ️ {contact['note']}"
        await update.message.reply_text(text, parse_mode="Markdown")
        return

    # Admin setting contact info
    if not context.args:
        contact = await db.get_admin_contact()
        current = ""
        if contact:
            current = (
                f"\n\n*Current Contact:*\n"
                f"Telegram: {contact.get('telegram') or 'Not set'}\n"
                f"Phone: {contact.get('phone') or 'Not set'}\n"
                f"Note: {contact.get('note') or 'Not set'}"
            )
        await update.message.reply_text(
            f"Usage: `/adcm <telegram> <phone> [note]`\n\n"
            f"Examples:\n"
            f"`/adcm @adminuser +91999999 Contact for OTP help`\n"
            f"`/adcm @adminuser none Support available 9am-9pm`{current}",
            parse_mode="Markdown"
        )
        return

    telegram = context.args[0] if len(context.args) >= 1 else None
    phone = context.args[1] if len(context.args) >= 2 else None
    note = " ".join(context.args[2:]) if len(context.args) > 2 else None

    # Handle "none" values
    if telegram and telegram.lower() == "none":
        telegram = None
    if phone and phone.lower() == "none":
        phone = None

    await db.set_admin_contact(telegram, phone, note)
    await update.message.reply_text(
        f"✅ *Admin contact updated!*\n\n"
        f"📱 Telegram: {telegram or 'Not set'}\n"
        f"☎️ Phone: {phone or 'Not set'}\n"
        f"ℹ️ Note: {note or 'Not set'}\n\n"
        f"Users will see this when they run /adcm or ask for OTP help.",
        parse_mode="Markdown"
    )


async def cmd_clist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Full command list with usage — admin section visible to admins only."""
    user_section = (
        "📜 *Full Command List*\n\n"
        "👤 *User Commands*\n\n"
        "`/start` — Welcome message\n"
        "`/getaccount` — Request an account\n"
        "  └ Bot will ask for your mobile number first\n"
        "`/myotp` — View your one-time OTP code\n"
        "  └ Can only be used once per account\n"
        "`/myaccount` — View your assigned account details\n"
        "`/done` — Signal you're done & request withdrawal\n"
        "  └ Bot shows wallet address to send to\n"
        "  └ Send a screenshot of your withdrawal\n"
        "`/adcm` — View admin contact info\n"
        "`/help` — Quick command overview\n"
        "`/clist` — This full command list"
    )

    admin_section = (
        "\n\n👑 *Admin Commands*\n\n"
        "📤 *Account Management*\n"
        "`/upload` — Upload Excel file with accounts\n"
        "  └ Send `.xlsx` after running this command\n"
        "  └ Columns: ID | Password | TOTP Secret | Status\n"
        "`/list` — View all accounts and their status\n"
        "`/stats` — Quick account count summary\n"
        "`/reset` — Mark ALL accounts as available\n"
        "`/resetaccount <id>` — Reset one specific account\n"
        "  └ Usage: `/resetaccount user@example.com`\n"
        "`/remove <id>` — Permanently delete an account\n"
        "  └ Usage: `/remove user@example.com`\n\n"
        "👥 *User Whitelist*\n"
        "`/uploadusers` — Upload Excel with allowed phone numbers\n"
        "  └ Send `.xlsx` after running — Columns: Phone | Name\n"
        "`/listusers` — View all whitelisted users\n"
        "`/adduser <phone> <name>` — Add a single user\n"
        "  └ Usage: `/adduser +919876543210 Rahul Sharma`\n"
        "`/removeuser <phone>` — Remove a user from whitelist\n"
        "  └ Usage: `/removeuser +919876543210`\n"
        "`/clearusers` — Remove ALL users from whitelist\n\n"
        "🔐 *OTP*\n"
        "`/adotp <account_id>` — Get OTP for any account\n"
        "  └ Usage: `/adotp user@example.com`\n\n"
        "💸 *Withdrawals*\n"
        "`/withdrawals` — View all pending withdrawal requests\n"
        "`/approve <id> [note]` — Approve withdrawal & free account\n"
        "  └ Usage: `/approve 5` or `/approve 5 Payment confirmed`\n"
        "`/reject <id> [reason]` — Reject a withdrawal request\n"
        "  └ Usage: `/reject 5 Invalid screenshot`\n\n"
        "💳 *Wallets*\n"
        "`/setwallet <address> <network> [label]` — Add wallet\n"
        "  └ Usage: `/setwallet TRxxxxxxx USDT_TRC20 MainWallet`\n"
        "`/wallets` — View all saved withdrawal wallets\n\n"
        "🔔 *Notifications*\n"
        "`/setgroup` — Set this group for withdrawal alerts\n"
        "  └ Run this command *inside* your admin group\n"
        "`/getgroup` — View current notification group\n\n"
        "⚙️ *Settings*\n"
        "`/adcm <telegram> <phone> [note]` — Set admin contact\n"
        "  └ Usage: `/adcm @admin +91999999 9am-9pm IST`\n\n"
        "👑 *Admin Management*\n"
        "`/addadmin <user_id>` — Add a new admin\n"
        "  └ Usage: `/addadmin 123456789`\n"
        "`/removeadmin <user_id>` — Remove a DB admin\n"
        "  └ Usage: `/removeadmin 123456789`\n"
        "`/admins` — List all current admins"
    )

    text = user_section
    if await is_admin_full(update.effective_user.id):
        text += admin_section

    await update.message.reply_text(text, parse_mode="Markdown")


async def cmd_setgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin sets the group for withdrawal notifications. Run this inside the group."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    chat = update.effective_chat
    if chat.type not in ("group", "supergroup"):
        await update.message.reply_text(
            "❌ Run this command *inside the group* you want notifications sent to.\n\n"
            "Steps:\n"
            "1. Add this bot to your admin group\n"
            "2. Run `/setgroup` inside that group",
            parse_mode="Markdown"
        )
        return

    await db.set_notify_group_id(chat.id)
    await update.message.reply_text(
        f"✅ *Notification group set!*\n\n"
        f"Group: *{chat.title}*\n"
        f"ID: `{chat.id}`\n\n"
        f"All withdrawal requests will now be sent here.",
        parse_mode="Markdown"
    )


async def cmd_getgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin checks the current notification group."""
    if not await is_admin_full(update.effective_user.id):
        await update.message.reply_text("❌ Admin only.")
        return

    group_id = await db.get_notify_group_chat_id()
    if not group_id:
        await update.message.reply_text(
            "⚠️ No notification group set.\n"
            "Go to your admin group and run `/setgroup`.",
            parse_mode="Markdown"
        )
        return

    await update.message.reply_text(
        f"✅ Notifications are sent to group ID: `{group_id}`\n\n"
        f"To change it, run `/setgroup` inside a different group.",
        parse_mode="Markdown"
    )


async def handle_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Route text input to the right handler based on user state."""
    if context.user_data.get("waiting_for_mobile"):
        await handle_mobile_number(update, context)
    elif context.user_data.get("waiting_for_txn"):
        # Legacy fallback — should not normally be reached
        pass


# --- App Setup ---

async def post_init(application: Application):
    await db.init()
    await db.run_migrations()

    # User-facing commands (shown to everyone)
    user_cmds = [
        BotCommand("start", "Welcome message"),
        BotCommand("getaccount", "Get an available account with OTP"),
        BotCommand("myotp", "View your one-time OTP code"),
        BotCommand("myaccount", "View your account details"),
        BotCommand("done", "Withdraw earnings & release account"),
        BotCommand("adcm", "Contact admin for help"),
        BotCommand("help", "Quick command overview"),
        BotCommand("clist", "Full command list with usage"),
    ]
    await application.bot.set_my_commands(user_cmds)

    # Admin commands (shown only in admin private chats)
    admin_cmds = user_cmds + [
        BotCommand("upload", "Upload Excel file with accounts"),
        BotCommand("list", "View all accounts and their status"),
        BotCommand("stats", "View account statistics"),
        BotCommand("reset", "Mark all accounts as available"),
        BotCommand("resetaccount", "Reset a specific account"),
        BotCommand("remove", "Remove an account permanently"),
        BotCommand("withdrawals", "View pending withdrawals"),
        BotCommand("approve", "Approve a withdrawal request"),
        BotCommand("reject", "Reject a withdrawal request"),
        BotCommand("setwallet", "Set withdrawal wallet address"),
        BotCommand("wallets", "View all withdrawal wallets"),
        BotCommand("setgroup", "Set group for withdrawal notifications"),
        BotCommand("getgroup", "View current notification group"),
        BotCommand("addadmin", "Add a new admin"),
        BotCommand("removeadmin", "Remove an admin"),
        BotCommand("admins", "List all admins"),
    ]

    from telegram import BotCommandScopeChat
    for admin_id in ADMIN_IDS:
        try:
            await application.bot.set_my_commands(
                admin_cmds,
                scope=BotCommandScopeChat(chat_id=admin_id)
            )
        except Exception as e:
            logger.warning(f"Could not set admin commands for {admin_id}: {e}")

    # Also set for DB admins
    db_admins = await db.get_admins()
    for a in db_admins:
        if a["user_id"] not in ADMIN_IDS:
            try:
                await application.bot.set_my_commands(
                    admin_cmds,
                    scope=BotCommandScopeChat(chat_id=a["user_id"])
                )
            except Exception as e:
                logger.warning(f"Could not set admin commands for DB admin {a['user_id']}: {e}")

    logger.info(f"Dashboard bot started. Admins: {ADMIN_IDS}")


def main():
    app = Application.builder().token(TOKEN).post_init(post_init).build()

    # User commands
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("clist", cmd_clist))
    app.add_handler(CommandHandler("getaccount", cmd_getaccount))
    app.add_handler(CommandHandler("myotp", cmd_myotp))
    app.add_handler(CommandHandler("myaccount", cmd_myaccount))
    app.add_handler(CommandHandler("done", cmd_done))
    app.add_handler(CommandHandler("adcm", cmd_adcm))

    # Admin commands
    app.add_handler(CommandHandler("upload", cmd_upload))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("reset", cmd_reset))
    app.add_handler(CommandHandler("resetaccount", cmd_resetaccount))
    app.add_handler(CommandHandler("remove", cmd_remove))
    app.add_handler(CommandHandler("adotp", cmd_adotp))
    app.add_handler(CommandHandler("uploadusers", cmd_uploadusers))
    app.add_handler(CommandHandler("listusers", cmd_listusers))
    app.add_handler(CommandHandler("adduser", cmd_adduser))
    app.add_handler(CommandHandler("removeuser", cmd_removeuser))
    app.add_handler(CommandHandler("clearusers", cmd_clearusers))
    app.add_handler(CommandHandler("addadmin", cmd_addadmin))
    app.add_handler(CommandHandler("removeadmin", cmd_removeadmin))
    app.add_handler(CommandHandler("admins", cmd_admins))
    app.add_handler(CommandHandler("approve", cmd_approve))
    app.add_handler(CommandHandler("reject", cmd_reject))
    app.add_handler(CommandHandler("withdrawals", cmd_withdrawals))
    app.add_handler(CommandHandler("setwallet", cmd_setwallet))
    app.add_handler(CommandHandler("wallets", cmd_wallets))
    app.add_handler(CommandHandler("setgroup", cmd_setgroup))
    app.add_handler(CommandHandler("getgroup", cmd_getgroup))

    # File upload handler
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    # Photo handler for withdrawal screenshots
    app.add_handler(MessageHandler(filters.PHOTO, handle_screenshot))

    # Text input handler (mobile number collection)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_input))

    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
