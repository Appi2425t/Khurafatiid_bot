"""Run this once to generate sample Excel templates: python create_sample.py"""
import openpyxl
from openpyxl.styles import Font, PatternFill

# ─── Accounts Excel ───────────────────────────────────────────────────────────
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Accounts"
ws1.append(["ID", "Password", "TOTP Secret", "Status"])

ws1.append(["user1@example.com", "Pass@1234", "JBSWY3DPEHPK3PXP", "available"])
ws1.append(["user2@example.com", "Admin@5678", "JBSWY3DPEHPK3PXQ", "available"])
ws1.append(["user3@example.com", "Test@9999", "JBSWY3DPEHPK3PXR", "available"])

for cell in ws1[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 35
ws1.column_dimensions["D"].width = 15

wb1.save("sample_accounts.xlsx")
print("✅ sample_accounts.xlsx created!")

# ─── Users / Whitelist Excel ──────────────────────────────────────────────────
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Users"
ws2.append(["Phone", "Name"])

ws2.append(["+91 9876543210", "Rahul Sharma"])
ws2.append(["+91 8765432109", "Priya Singh"])
ws2.append(["+91 7654321098", "Amit Kumar"])
ws2.append(["+91 6543210987", "Sneha Patel"])

for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1e7e34", end_color="1e7e34", fill_type="solid")

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 30

wb2.save("sample_users.xlsx")
print("✅ sample_users.xlsx created!")
print("\nDone! Upload these files using /upload and /uploadusers in the bot.")
